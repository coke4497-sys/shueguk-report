# 슈국 수파베이스 일일 점검·복구 도구
#
# 사용법:
#   python3 audit_heal.py <리포트시트.xlsx> [--omr <OMR시트.xlsx>] [--hwork <HWORK시트.xlsx>]
#                         [--voca <어휘시트.xlsx>] [--signup] [--clinic] [--heal]
#     --omr    : 주말 모의고사 OMR 시트(회차정답·응답)도 대조·동기화
#     --hwork  : H WORK 시트(HWORK목록·제출기록)도 대조·동기화
#     --voca   : 어휘 결과 시트(첫 탭)도 대조·동기화
#     --signup : 주말 신청 데이터를 신청 백엔드 API(action=data)에서 받아 대조·동기화
#     --clinic : 클리닉 신청·설정을 클리닉 백엔드 API(action=data)에서 받아 대조·동기화
#
# 하는 일:
#   · 시트(원본/미러 각 방향)와 수파베이스를 표 단위로 전수 대조해 요약을 출력한다.
#   · --heal 을 주면 안전한 방향으로만 자동 복구한다:
#       - 시트가 원본인 표(시간표·기간·학생정보·공지·공지확인·별·시험·문항):
#         시트 내용으로 수파베이스를 맞춘다(갱신·추가·삭제).
#       - 수파베이스가 원본인 표(출석·메모·숙제검사·내신기록·제출결과):
#         시트에만 있는 행(옛 페이지·미러 지연으로 생긴 것)을 수파베이스에 추가만 한다.
#         수파베이스 행을 시트 값으로 덮어쓰지 않는다. 시트에 없는 수파베이스 행은
#         이중 기록 실패이므로 개수만 보고한다(시트 쓰기는 이 도구 밖).
#       - 지필일정: 수파베이스가 원본(시트 탭 동결) — 보고만.
#   · 이동기록은 양쪽 다 쓰므로 서로 없는 행을 수파베이스에 추가만 한다.
#
# 키(비밀 아님 — 페이지에 공개된 publishable 키)는 아래 상수 사용.
import sys, os, json, re, ssl, datetime, urllib.request

SB_URL = 'https://bangdbhqpphqqdwcledg.supabase.co/rest/v1'
SB_KEY = 'sb_publishable_dE9d1KIbpgYaQkaS2MSrlg_-7SiRJuT'

def _ctx():
    ca = os.environ.get('SSL_CERT_FILE') or os.environ.get('CURL_CA_BUNDLE') or '/root/.ccr/ca-bundle.crt'
    return ssl.create_default_context(cafile=ca if os.path.exists(ca) else None)
CTX = _ctx()

def sb(method, path, body=None, prefer=None):
    req = urllib.request.Request(SB_URL + path, method=method,
        data=json.dumps(body).encode() if body is not None else None,
        headers={'apikey': SB_KEY, 'Authorization': 'Bearer ' + SB_KEY,
                 'Content-Type': 'application/json', **({'Prefer': prefer} if prefer else {})})
    with urllib.request.urlopen(req, context=CTX) as r:
        t = r.read()
        return json.loads(t) if t else None

def sb_all(table, params=''):
    out, frm = [], 0
    while True:
        req = urllib.request.Request(f'{SB_URL}/{table}?{params}' if params else f'{SB_URL}/{table}',
            headers={'apikey': SB_KEY, 'Authorization': 'Bearer ' + SB_KEY,
                     'Range-Unit': 'items', 'Range': f'{frm}-{frm+999}'})
        with urllib.request.urlopen(req, context=CTX) as r:
            rows = json.loads(r.read())
        out += rows
        if len(rows) < 1000: return out
        frm += 1000

# ── 시트 추출 ─────────────────────────────────────────────
import openpyxl
def txt(v):
    if v is None: return ''
    if isinstance(v, float) and v == int(v): return str(int(v))
    return str(v).strip()
def ts(v, dflt='2026-01-01T00:00:00+09:00'):
    if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%dT%H:%M:%S+09:00')
    s = txt(v)
    return s if s else dflt
def ymd(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime('%Y-%m-%d')
    return txt(v)
def hm(v):
    if isinstance(v, datetime.time): return v.strftime('%H:%M')
    if isinstance(v, datetime.datetime): return v.strftime('%H:%M')
    return txt(v)
def text_restore(v):
    if isinstance(v, datetime.datetime):
        dow = '일월화수목금토'[(v.weekday() + 1) % 7]
        base = f'{v.month}/{v.day} {dow}'
        if v.hour == 0 and v.minute == 0: return base
        h12 = 12 if v.hour % 12 == 0 else v.hour % 12
        return f'{base} {h12}:{v.minute:02d}'
    if isinstance(v, datetime.date):
        dow = '일월화수목금토'[(v.weekday() + 1) % 7]
        return f'{v.month}/{v.day} {dow}'
    return txt(v)
def plain(n): return re.sub(r'\(.*\)$', '', str(n or '')).strip()
def ep(s):
    try: return round(datetime.datetime.fromisoformat(s).timestamp())
    except Exception: return None

def extract(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    def rows(name):
        if name not in wb.sheetnames: return []
        it = wb[name].iter_rows(values_only=True); next(it, None)
        return [r for r in it if r and any(c is not None and str(c).strip() != '' for c in r)]
    def col(r, i): return r[i] if len(r) > i else None
    def book(v): return '내신' if txt(v) == '내신' else '정규'
    d = {}
    d['tt_classes'] = [{'book': bk, 'class_id': txt(col(r,0)), 'day': txt(col(r,1)),
                        'start_time': hm(col(r,2)), 'end_time': hm(col(r,3)), 'location': txt(col(r,4)),
                        'teacher': txt(col(r,5)), 'name': txt(col(r,6)), 'roster': txt(col(r,7))}
                       for bk, tab in [('정규','정규시간표'),('내신','내신시간표')] for r in rows(tab) if txt(col(r,0))]
    att = {}
    for r in rows('출석기록'):
        dt = ymd(col(r,0)); st = txt(col(r,4))
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', dt) or st not in ('출석','지각','결석'): continue
        rec = {'date': dt, 'book': book(col(r,1)), 'class_id': txt(col(r,2)), 'student': txt(col(r,3)),
               'status': st, 'memo': txt(col(r,5)), 'makeup_plan': txt(col(r,7)), 'makeup_done': txt(col(r,8)),
               'makeup_memo': txt(col(r,9)), 'updated_at': ts(col(r,6))}
        att[(dt, rec['book'], rec['class_id'], plain(rec['student']))] = rec
    d['attendance'] = list(att.values())
    d['tt_log'] = [{'at': ts(col(r,0)), 'kind': txt(col(r,1)), 'student': txt(col(r,2)),
                    'from_class_id': txt(col(r,3)), 'from_class': txt(col(r,4)), 'to_class_id': txt(col(r,5)),
                    'to_class': txt(col(r,6)), 'reason': txt(col(r,7)), 'book': book(col(r,8)),
                    'apply_date': (lambda a: a if re.match(r'^\d{4}-\d{2}-\d{2}$', a) else None)(ymd(col(r,9)))}
                   for r in rows('시간표이동기록') if txt(col(r,1)) or txt(col(r,2))]
    d['tt_memo'] = [{'date': ymd(col(r,0)), 'memo': txt(col(r,1))} for r in rows('시간표메모')
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', ymd(col(r,0)))]
    d['tt_period'] = [{'week_wednesday': ymd(col(r,0)), 'book': book(col(r,1))} for r in rows('기간설정')
                      if re.match(r'^\d{4}-\d{2}-\d{2}$', ymd(col(r,0)))]
    hw = {}
    for r in rows('숙제검사'):
        wk = ymd(col(r,1)); tok = txt(col(r,2))
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', wk) or not tok: continue
        try: scores = json.loads(txt(col(r,6)) or '{}')
        except Exception: scores = {}
        hw[(wk, tok)] = {'week': wk, 'token': tok, 'name': txt(col(r,3)), 'school': txt(col(r,4)),
                         'grade': txt(col(r,5)), 'scores': scores, 'max': int(col(r,7) or 0), 'pct': int(col(r,8) or 0),
                         'pub': text_restore(col(r,9)), 'priv': text_restore(col(r,10)),
                         'missing': txt(col(r,11)) == '미제출', 'plan': text_restore(col(r,12)),
                         'plan_done': txt(col(r,13)) == '완료', 'at': ts(col(r,0))}
    d['hwcheck_records'] = list(hw.values())
    ns = {}
    for r in rows('내신기록'):
        p, ck, kd = txt(col(r,0)), txt(col(r,1)), txt(col(r,2))
        if not p or not ck or kd not in ('범위','주차','학생','클리어'): continue
        wk = ymd(col(r,3)) if kd in ('주차','학생') else ''
        stn = txt(col(r,4)) if kd == '학생' else ''
        ns[(p, ck, kd, wk, stn)] = {'period': p, 'class_key': ck, 'kind': kd, 'week': wk, 'student': stn,
                                    'text1': txt(col(r,5)), 'text2': txt(col(r,6)), 'at': ts(col(r,7))}
    d['naeshin_records'] = list(ns.values())
    d['exams'] = [{'report_id': txt(col(r,0)), 'title': txt(col(r,1)), 'review': txt(col(r,2)),
                   'scope': txt(col(r,3)), 'school': txt(col(r,4)), 'grade': txt(col(r,5))}
                  for r in rows('보고서목록') if txt(col(r,0))]
    qs, seq = [], 0
    for r in rows('문항'):
        if not txt(col(r,0)): continue
        seq += 1
        m = txt(col(r,8)).lower()
        qs.append({'report_id': txt(col(r,0)), 'seq': seq, 'no': txt(col(r,1)), 'area': txt(col(r,2)),
                   'qtype': txt(col(r,3)), 'lv': txt(col(r,4)), 'txt': txt(col(r,5)), 'detail': txt(col(r,6)),
                   'grp': txt(col(r,7)), 'multi': m in ('1','y','yes','true','o') or m.startswith('복수')})
    d['exam_questions'] = qs
    d['submissions'] = [{'submitted_at': ts(col(r,0)), 'exam': txt(col(r,1)), 'school': txt(col(r,2)),
                         'grade': txt(col(r,3)), 'name': txt(col(r,4)), 'wrong_count': txt(col(r,5)),
                         'wrong_text': txt(col(r,6)), 'vow': txt(col(r,7)), 'teacher_note': txt(col(r,8)),
                         'score': txt(col(r,9)), 'parent_phone': txt(col(r,10))}
                        for r in rows('제출결과') if txt(col(r,4))]
    stu, seq = [], 0
    for r in rows('학생정보'):
        if not txt(col(r,1)): continue
        seq += 1
        stu.append({'seq': seq, 'student_id': txt(col(r,0)), 'name': txt(col(r,1)), 'school': txt(col(r,2)),
                    'grade': txt(col(r,3)), 'teacher': txt(col(r,4)), 'memo': txt(col(r,5)),
                    'class_a': txt(col(r,6)), 'class_b': txt(col(r,7)), 'naeshin_a': txt(col(r,8)),
                    'naeshin_b': txt(col(r,9)), 'enrolled': txt(col(r,10)), 'code': txt(col(r,11)),
                    'student_phone': txt(col(r,12)), 'reg_date': ymd(col(r,13))})
    d['students'] = stu
    nt, seq = [], 0
    for r in rows('공지'):
        title, body, target, typ = txt(col(r,3)), txt(col(r,4)), txt(col(r,2)), txt(col(r,1))
        if not (title or body or target or typ): continue
        seq += 1
        nt.append({'seq': seq, 'date': ymd(col(r,0)), 'type': typ, 'target': target, 'title': title,
                   'body': body, 'hidden': bool(re.match(r'^(n|숨김|off|x)$', txt(col(r,5)), re.I))})
    d['notices'] = nt
    d['notice_reads'] = [{'at': ts(col(r,0)), 'student_id': txt(col(r,1)), 'name': txt(col(r,2)),
                          'school': txt(col(r,3)), 'notice_key': txt(col(r,4))} for r in rows('공지확인')]
    d['star_bonus'] = [{'at': ts(col(r,0)), 'student_id': txt(col(r,1)), 'name': txt(col(r,2)),
                        'school': txt(col(r,3)), 'stars': int(col(r,4) or 0), 'reason': txt(col(r,5)),
                        'grade': txt(col(r,6))} for r in rows('별') if txt(col(r,2))]
    return d

def extract_omr(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    exams = []
    for r in wb['회차정답'].iter_rows(values_only=True):   # 헤더 없음 — 1행부터 데이터
        if not r or not txt(r[0]): continue
        data = txt(r[1] if len(r) > 1 else '')
        exams.append({'name': txt(r[0]), 'data': data,
                      'mode': '통합' if re.search(r'"mode"\s*:\s*"통합"', data) else ''})
    rows = list(wb['응답'].iter_rows(values_only=True))
    hdr = [txt(c) for c in rows[0]] if rows else []
    idcol = hdr.index('학생ID') if '학생ID' in hdr else -1
    resp = []
    for r in rows[1:]:
        if not r or not txt(r[3] if len(r) > 3 else None): continue
        ans = {}
        for q in range(1, 46):
            v = txt(r[9+q] if len(r) > 9+q else None)
            if v: ans[str(q)] = v
        dt = r[2]
        dk = f'{dt.month}월 {dt.day}일' if isinstance(dt, (datetime.datetime, datetime.date)) else txt(dt)
        resp.append({'submitted_at': ts(r[0]), 'exam': txt(r[1]), 'exam_date': dk,
                     'name': txt(r[3]), 'school': txt(r[4]), 'grade': txt(r[5]), 'subject': txt(r[6]),
                     'got': txt(r[7]), 'total': txt(r[8]), 'level': txt(r[9]), 'answers': ans,
                     'student_id': txt(r[idcol] if idcol >= 0 and len(r) > idcol else None)})
    return {'omr_exams': exams, 'omr_responses': resp}

SIGNUP_EXEC = 'https://script.google.com/macros/s/AKfycbzdqac0xTnCaOo_t_2swJQqdfxjiA14sTo-ThTV8VvwcwaTucM1MQGeJfMfV4lNLM75/exec'
def fetch_signup():
    # 신청 백엔드의 교사용 목록 API — 시트 원본 그대로 (pw는 공개 페이지 내장과 동일 수준)
    import urllib.parse
    KST = datetime.timezone(datetime.timedelta(hours=9))
    def ts_norm(v):
        s = txt(v)
        if not s: return ''
        if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}', s): return s[:16]
        try: return datetime.datetime.fromisoformat(s.replace('Z', '+00:00')).astimezone(KST).strftime('%Y-%m-%d %H:%M')
        except Exception: return s
    def date_norm(v):
        s = txt(v)
        if not s or ('월' in s and '일' in s): return s
        try:
            d = datetime.datetime.fromisoformat(s.replace('Z', '+00:00')).astimezone(KST)
            return f'{d.month}월 {d.day}일'
        except Exception: return s
    url = SIGNUP_EXEC + '?action=data&pw=sh'
    try:
        req = urllib.request.Request(url)
        for _ in range(5):
            with urllib.request.urlopen(req, context=CTX) as r:
                if r.status in (301, 302, 303, 307):
                    req = urllib.request.Request(r.headers['Location']); continue
                d = json.loads(r.read()); break
        if d.get('result') != 'success': return None
    except Exception:
        return None
    return [{'ts': ts_norm(r.get('제출시각')), 'name': txt(r.get('이름')), 'school': txt(r.get('학교')),
             'grade': txt(r.get('학년')), 'student_id': txt(r.get('학생ID')), 'subject': txt(r.get('선택과목')),
             'day': txt(r.get('응시요일')), 'exam_date': date_norm(r.get('응시일자'))} for r in d.get('rows', [])]

def extract_hwork(xlsx):
    # H WORK 시트 — HWORK목록(A강사 B제목 C데이터JSON D마감일)·제출기록(11열)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    def rows(name):
        if name not in wb.sheetnames: return []
        it = wb[name].iter_rows(values_only=True); next(it, None)
        return [r for r in it if r and any(c is not None and str(c).strip() != '' for c in r)]
    def col(r, i): return r[i] if len(r) > i else None
    def jload(v):
        try: return json.loads(txt(v) or '{}')
        except Exception: return {}
    def tsmin(v):
        if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%d %H:%M')
        s = txt(v)
        return s[:16] if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}', s) else s
    def num(v):
        try: return int(float(txt(v) or 0))
        except Exception: return 0
    hw = [{'teacher': txt(col(r,0)), 'code': txt(col(r,1)), 'data': jload(col(r,2)), 'due': ymd(col(r,3))}
          for r in rows('HWORK목록') if txt(col(r,0)) and txt(col(r,1))]
    subs = [{'ts': tsmin(col(r,0)), 'teacher': txt(col(r,1)), 'code': txt(col(r,2)),
             'school': txt(col(r,3)), 'grade': txt(col(r,4)), 'name': txt(col(r,5)),
             'got': num(col(r,6)), 'total': num(col(r,7)), 'answers': jload(col(r,8)),
             'detail': jload(col(r,9)), 'questions': jload(col(r,10))}
            for r in rows('제출기록') if txt(col(r,1)) and txt(col(r,2))]
    return {'hwork_homeworks': hw, 'hwork_submissions': subs}

def extract_voca(xlsx):
    # 어휘 결과 시트 첫 탭 — 실제 사용 열: A타임스탬프 B이름 C학교 D학년 F점수 I phone4 J round K details
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True); next(it, None)
    def tsmin(v):
        if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%d %H:%M')
        s = txt(v)
        return s[:16] if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}', s) else s
    def col(r, i): return r[i] if len(r) > i else None
    return [{'ts': tsmin(col(r,0)), 'name': txt(col(r,1)), 'school': txt(col(r,2)), 'grade': txt(col(r,3)),
             'score': txt(col(r,5)), 'phone4': txt(col(r,8)), 'round': txt(col(r,9)), 'details': txt(col(r,10))}
            for r in it if r and txt(col(r,1))]

CLINIC_EXEC = 'https://script.google.com/macros/s/AKfycbw8e-054e4VUfRRx-PadyuoXRb-jxsKRcdOaq04SJH1oJyFVS_VOh9GUN_pL5cHPPzKVA/exec'
def fetch_clinic():
    # 클리닉 백엔드의 교사용 목록 API — '응답' 시트 + Script Properties 설정 (pw는 공개 페이지 내장과 동일 수준)
    KST = datetime.timezone(datetime.timedelta(hours=9))
    def ts_norm(v):
        s = txt(v)
        if not s: return ''
        if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}', s): return s[:16]
        try: return datetime.datetime.fromisoformat(s.replace('Z', '+00:00')).astimezone(KST).strftime('%Y-%m-%d %H:%M')
        except Exception: return s
    url = CLINIC_EXEC + '?action=data&pw=sh'
    try:
        req = urllib.request.Request(url)
        for _ in range(5):
            with urllib.request.urlopen(req, context=CTX) as r:
                if r.status in (301, 302, 303, 307):
                    req = urllib.request.Request(r.headers['Location']); continue
                d = json.loads(r.read()); break
        if d.get('result') != 'success': return None
    except Exception:
        return None
    rows = [{'ts': ts_norm(r.get('제출시각')), 'name': txt(r.get('이름')), 'school': txt(r.get('학교')),
             'phone': txt(r.get('전화뒤4')), 'slot': txt(r.get('클리닉시간')), 'rtype': txt(r.get('유형')),
             'area': txt(r.get('영역')), 'content': txt(r.get('구체내용')), 'qcount': txt(r.get('질문개수')),
             'memo': txt(r.get('메모')), 'grade': txt(r.get('학년')), 'student_id': txt(r.get('학생ID')),
             'teacher': txt(r.get('담당강사')), 'token': txt(r.get('토큰')), 'clear': ts_norm(r.get('클리어'))}
            for r in d.get('rows', [])
            if txt(r.get('이름')) or txt(r.get('유형')) or txt(r.get('영역')) or txt(r.get('구체내용'))]
    settings = {k: d.get(k) for k in ('open', 'slots', 'allSlots', 'teachers', 'target') if k in d}
    return rows, settings

# ── 대조·복구 ─────────────────────────────────────────────
ISSUES = []
def report(line): ISSUES.append(line); print('  !', line)

def sync_sheet_master(table, sheet_rows, keyf, valf, heal, id_field='id'):
    """시트가 원본 — 수파베이스를 시트에 맞춘다."""
    cur = sb_all(table)
    curmap = {}
    for r in cur: curmap.setdefault(keyf(r), []).append(r)
    smap = {}
    for r in sheet_rows: smap[keyf(r)] = r
    to_add = [r for k, r in smap.items() if k not in curmap]
    to_del = [r for k, rs in curmap.items() if k not in smap for r in rs]
    to_upd = [(r, curmap[k][0]) for k, r in smap.items() if k in curmap and valf(r) != valf(curmap[k][0])]
    n = len(to_add) + len(to_del) + len(to_upd)
    tag = f'{table}: 추가 {len(to_add)} · 갱신 {len(to_upd)} · 삭제 {len(to_del)}'
    if n == 0:
        print(f'  = {table}: 일치 ({len(sheet_rows)}행)')
        return
    report(tag)
    if not heal: return
    for r in to_add: sb('POST', f'/{table}', [r], 'return=minimal')
    for r, c in to_upd: sb('PATCH', f'/{table}?{id_field}=eq.{c[id_field]}', r)
    for c in to_del: sb('DELETE', f'/{table}?{id_field}=eq.{c[id_field]}')
    print(f'    → 복구 완료')

def insert_missing(table, sheet_rows, keyf, cur_keyf, heal, label):
    """수파베이스가 원본(또는 양쪽 기록) — 시트에만 있는 행을 추가만."""
    cur = sb_all(table)
    have = {cur_keyf(r) for r in cur}
    missing = [r for r in sheet_rows if keyf(r) not in have]
    sheet_keys = {keyf(r) for r in sheet_rows}
    db_only = [r for r in cur if cur_keyf(r) not in sheet_keys]
    if not missing and not db_only:
        print(f'  = {table}: 일치 ({len(cur)}행)')
        return
    if missing:
        report(f'{table}: 시트에만 있는 행 {len(missing)} — DB에 추가{"" if heal else " 필요"}')
        if heal:
            for r in missing: sb('POST', f'/{table}', [r], 'return=minimal')
            print('    → 추가 완료')
    if db_only:
        report(f'{table}: DB에만 있는 행 {len(db_only)} — {label}')

def main():
    xlsx = sys.argv[1]
    heal = '--heal' in sys.argv
    d = extract(xlsx)
    print(f'[audit] {datetime.datetime.now().isoformat()} heal={heal}')

    print('· 시트가 원본인 표')
    sync_sheet_master('tt_classes', d['tt_classes'], lambda r: (r['book'], r['class_id']),
        lambda r: (r['day'], r['start_time'], r['end_time'], r['location'], r['teacher'], r['name'], r['roster']), heal)
    sync_sheet_master('tt_period', d['tt_period'], lambda r: r['week_wednesday'], lambda r: r['book'], heal, 'week_wednesday')
    sync_sheet_master('students', d['students'], lambda r: (r['student_id'], r['name']),
        lambda r: (r['school'], r['grade'], r['teacher'], r['memo'], r['class_a'], r['class_b'],
                   r['naeshin_a'], r['naeshin_b'], r['enrolled'], r['code'], r['student_phone'], r['reg_date']), heal)
    sync_sheet_master('notices', d['notices'], lambda r: (r['title'], r['date'], r['type'], r['target']),
        lambda r: (r['body'], r['hidden']), heal)
    sync_sheet_master('notice_reads', d['notice_reads'],
        lambda r: (ep(r['at']), r.get('student_id',''), r.get('notice_key','')),
        lambda r: r.get('name',''), heal)
    sync_sheet_master('star_bonus', d['star_bonus'],
        lambda r: (ep(r['at']), r.get('name',''), r.get('stars',0)), lambda r: r.get('reason',''), heal)
    sync_sheet_master('exams', d['exams'], lambda r: r['report_id'],
        lambda r: (r['title'], r['review'], r['scope'], r['school'], r['grade']), heal, 'report_id')
    sync_sheet_master('exam_questions', d['exam_questions'], lambda r: (r['report_id'], r['seq']),
        lambda r: (r['no'], r['area'], r['qtype'], r['lv'], r['txt'], r['detail'], r['grp'], r['multi']), heal)

    print('· 수파베이스가 원본인 표 (시트에만 있는 행 추가 / DB에만 있는 행은 보고)')
    insert_missing('attendance', d['attendance'],
        lambda r: (r['date'], r['book'], r['class_id'], plain(r['student'])),
        lambda r: (r['date'], r['book'], r['class_id'], r['student_plain']), heal,
        '이중 기록 지연·실패 가능 — 다음 점검에도 남으면 확인')
    insert_missing('tt_memo', d['tt_memo'], lambda r: r['date'], lambda r: r['date'], heal, '시트 미러 지연')
    insert_missing('hwcheck_records', d['hwcheck_records'], lambda r: (r['week'], r['token']),
        lambda r: (r['week'], r['token']), heal, '이중 기록 지연·실패 가능')
    insert_missing('naeshin_records', d['naeshin_records'],
        lambda r: (r['period'], r['class_key'], r['kind'], r['week'], r['student']),
        lambda r: (r['period'], r['class_key'], r['kind'], r['week'], r['student']), heal, '이중 기록 지연·실패 가능')
    insert_missing('submissions', d['submissions'],
        lambda r: (ep(r['submitted_at']), r['name'].strip(), r['exam'].strip()),
        lambda r: (ep(r['submitted_at']), r['name'].strip(), r['exam'].strip()), heal, '이중 기록 지연·실패 가능')
    insert_missing('tt_log', d['tt_log'],
        lambda r: (ep(r['at']), r['kind'], r['student'], r['from_class_id'], r['to_class_id']),
        lambda r: (ep(r['at']), r['kind'], r['student'], r['from_class_id'], r['to_class_id']), heal,
        '페이지에서 직접 기록한 이동(1회 취소 등 정상일 수 있음)')

    # ── 주말 모의고사 (선택) — 셋 다 원본이 밖(시트/신청 백엔드)에 있는 미러 ──
    if '--omr' in sys.argv:
        omr_x = sys.argv[sys.argv.index('--omr') + 1]
        o = extract_omr(omr_x)
        print('· 주말 모의고사 OMR (시트가 원본)')
        sync_sheet_master('omr_exams', o['omr_exams'], lambda r: r['name'],
            lambda r: (r['data'], r['mode']), heal, 'name')
        sync_sheet_master('omr_responses', o['omr_responses'],
            lambda r: (ep(r['submitted_at']), r['name'], r['exam']),
            lambda r: (r['exam_date'], r['school'], r['grade'], r['subject'], r['got'], r['total'],
                       r['level'], json.dumps(r['answers'], sort_keys=True), r['student_id']), heal)
    if '--hwork' in sys.argv:
        hx = sys.argv[sys.argv.index('--hwork') + 1]
        h = extract_hwork(hx)
        print('· H WORK (시트가 원본)')
        sync_sheet_master('hwork_homeworks', h['hwork_homeworks'],
            lambda r: (r['teacher'], r['code']),
            lambda r: (json.dumps(r['data'], sort_keys=True), r['due']), heal)
        sync_sheet_master('hwork_submissions', h['hwork_submissions'],
            lambda r: (r['ts'], r['name'], r['code'], json.dumps(r['answers'], sort_keys=True)),
            lambda r: (r['teacher'], r['school'], r['grade'], r['got'], r['total'],
                       json.dumps(r['detail'], sort_keys=True), json.dumps(r['questions'], sort_keys=True)), heal)
    if '--voca' in sys.argv:
        vx = sys.argv[sys.argv.index('--voca') + 1]
        v = extract_voca(vx)
        print('· 어휘 테스트 (시트가 원본)')
        # 키에 점수·상세 포함 — 같은 분에 재제출한 중복 행(전송 재시도·재응시)이 서로 다른 키가 되도록
        sync_sheet_master('voca_results', v,
            lambda r: (r['ts'], r['name'], r['round'], r['score'], r['details']),
            lambda r: (r['school'], r['grade'], r['phone4']), heal)
    if '--signup' in sys.argv:
        rows = fetch_signup()
        if rows is None:
            report('signup_entries: 신청 백엔드 조회 실패 — 대조 못 함')
        else:
            print('· 주말 신청 (신청 시트가 원본, API로 대조)')
            sync_sheet_master('signup_entries', rows,
                lambda r: (r['ts'], r['name'], r['day'], r['student_id']),
                lambda r: (r['school'], r['grade'], r['subject'], r['exam_date']), heal)
    if '--clinic' in sys.argv:
        res = fetch_clinic()
        if res is None:
            report('clinic_requests: 클리닉 백엔드 조회 실패 — 대조 못 함')
        else:
            c_rows, c_set = res
            print('· 클리닉 (클리닉 시트·백엔드 설정이 원본, API로 대조)')
            sync_sheet_master('clinic_requests', c_rows,
                lambda r: (r['ts'], r['name'], r['phone'], r['slot'], r['rtype'], r['area'], r['content']),
                lambda r: (r['school'], r['qcount'], r['memo'], r['grade'], r['student_id'],
                           r['teacher'], r['token'], r['clear']), heal)
            cur_set = {r['key']: r['value'] for r in sb_all('clinic_settings')}
            diff = [k for k, v in c_set.items() if cur_set.get(k) != v]
            if not diff:
                print(f'  = clinic_settings: 일치 ({len(c_set)}키)')
            else:
                report(f'clinic_settings: 불일치 {", ".join(diff)}')
                if heal:
                    sb('POST', '/clinic_settings?on_conflict=key',
                       [{'key': k, 'value': c_set[k]} for k in diff],
                       'resolution=merge-duplicates,return=minimal')
                    print('    → 복구 완료')

    print()
    if ISSUES:
        print(f'[요약] 발견 {len(ISSUES)}건' + (' — 복구 시도함' if heal else ''))
        for x in ISSUES: print(' -', x)
        sys.exit(2 if not heal else 0)
    print('[요약] 전부 일치 — 이상 없음')

if __name__ == '__main__':
    main()
